import pathlib
import subprocess
import os
import openpyxl as xl
import pandas as pd
import win32print
import win32api
from itertools import islice
from tkinter.filedialog import askdirectory
import pygetwindow as gw
from datetime import datetime

"""
wb = xl.load_workbook('Controle RA - Copia.xlsm', data_only=False, keep_vba=True)
sheet = wb['ControleRA']
"""


class Agronomo:
    def __init__(self, instancia, nomeagr, cpf, art, ultimo, datamin, datamax, ramax):
        self.instancia = instancia
        self.nome = nomeagr
        self.cpf = cpf
        self.art = art
        self.ultimo = ultimo
        self.datamin = datamin
        self.datamax = datamax
        self.ramax = ramax

    def __str__(self) -> str:
        return 'Nome: %s CPF: %s' % (self.nome, self.cpf)


def abre_sol():
    path_sol = r'C:\orant\BIN\ifrun60.EXE'
    path_mec = r'J:\SISTEMAS\PIMSCS\EXE\APT_ATRC.EXE'
    path_plativ = r'J:\SISTEMAS\PIMSCS\EXE\PLATIV.EXE'
    path_loc_prod = r'J:\SISTEMAS\PIMSCS\EXE\LOCPROD.exe'
    path_config = r'J:\SISTEMAS\PIMSCS\EXE\CFGCEN.exe'

    # subprocess.Popen(path_mec, shell=False)

    titulo_tela = 'Sem título - Bloco de Notas'

    tela = gw.getWindowsWithTitle(titulo_tela)[0]

    # tela.activate()
    tela.restore()


def teste():
    """
    filename = askdirectory()
    print(filename)
    os.system(r"md C:\Temp\DataLoadRA")
    """

    i = 3

    wb = xl.load_workbook(filename=r'Controle RA.xlsm', data_only=True, keep_vba=True)
    cadastro = wb['ControleART']
    sheet = wb['ControleRA']

    agronomos = []
    i = 2
    while True:
        if cadastro['J' + str(i)].value is not None:
            agronomos.append(
                Agronomo(
                    cadastro['J' + str(i)].value,  # Instancia
                    cadastro['K' + str(i)].value,  # Nome
                    cadastro['L' + str(i)].value,  # CPF
                    cadastro['M' + str(i)].value,  # ART
                    cadastro['N' + str(i)].value,  # Ultima RA
                    str(cadastro['O' + str(i)].value.strftime("%d/%m/%Y")),  # Data Min
                    str(cadastro['P' + str(i)].value.strftime("%d/%m/%Y")),  # Data Max
                    cadastro['Q' + str(i)].value  # RA Max
                )
            )
            i = i + 1
        else:
            break

    for agr in agronomos:
        print(agr)


def teste_excel():
    wb = xl.load_workbook(filename=r'teste.xlsx', data_only=True, keep_vba=True, keep_links=True)
    wb.save(filename='teste.xlsm')


def teste_hora():
    from datetime import datetime
    while True:
        if int(datetime.today().strftime("%H")) < 23:  # Se a data for menor que 23, não faz nada
            print('não faz nada')
        elif int(datetime.today().strftime("%H")) > 2:  # Se a data for maior que 2 executa função.
            print('executa função')
            break


def arredondar(numero):
    if type(numero) == str:
        numero = int(numero)

    quociente, resto = numero // 5, numero % 5
    resultado = (5 * (quociente + (resto != 0)))
    if resultado == 0:
        resultado = 1
    return resultado


if __name__ == "__main__":
    print(arredondar('1.2'))
